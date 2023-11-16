import re
import os
import argparse
import openpyxl
from openpyxl.styles.borders import Border, Side

def check_and_create_dir(path):
    """
    check dir and create new dir
    """
    try:
        os.makedirs(path)
    except FileExistsError:
        return False

def parse_log_name(log_file):
    """
    parse log's name and return precison and core values
    support precison: bfloat16, woq_int8, static_int8
    """
    if re.search("bfloat16", log_file):
        precision=re.findall('bfloat16', log_file)[0]
        try:
            core=re.findall('\d{2}', log_file)[1]
        except IndexError:
            core='56'
    if re.search("woq_int8", log_file):
        precision=re.findall('woq_int8', log_file)[0]
        print(precision)
        try:
            core=re.findall('\d{2}', log_file)[0]
        except IndexError:
            core='56'
    if re.search("static_int8", log_file):
        precision=re.findall('static_int8', log_file)[0]
        try:
            core=re.findall('\d{2}', log_file)[0]
        except IndexError:
            core='56'
    return precision, core

def parse_to_excel(platform, report_path, log_file_name, loop, frequencys):
    """
    parse log to excel
    """
    precision, core= parse_log_name(log_file_name)
    # Workbook is created
    print(core)
    wb = openpyxl.Workbook()
    sheet = wb.active 
    sheet.title =core
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    row=1
    column=2
    fre_row=1
    fre_col = 1
    logfile_path = os.path.join(report_path,log_file_name)
    #Read output log file
    with open(logfile_path, 'r') as ds_log:
        lines = ds_log.readlines()
        for line in lines:
            # Get latency
            if re.search("Inference latency:", line):
                latency=re.findall('\d+\.\d+', line)[0]
                print(float(latency))
                cell = sheet.cell(row = row, column = column)
                cell.value = float(latency)
                cell.border = thin_border
                row+=1
            # Get first_token
            if re.search("First token average latency:", line):
                first_token=re.findall('\d+\.\d+', line)[0]
                print(float(first_token))
                cell = sheet.cell(row = row, column = column)
                cell.value = float(first_token)
                cell.border = thin_border
                row+=1
            # Get second_token
            if re.search("Average 2... latency:", line):
                second_token=re.findall('\d+\.\d+', line)[0]
                print(float(second_token))
                cell = sheet.cell(row = row, column = column)
                cell.value = float(second_token)
                cell.border = thin_border
                row+=1
            if re.search("WSF Portal URL:", line):
                dashboard_link = re.findall('https://.*', line)[0]
                sheet.cell(row=row-3, column=column).hyperlink = dashboard_link
                print(dashboard_link)
            # Get loop times
            if re.search("Current\s+.*\{[0-9]\}.*", line):
                loop_time=re.findall('([0-9])', line)[0]
                print(loop_time)
                current_frequency=re.findall('\d.\dGhz', line)[0]
                for frequency in frequencys:
                    if int(loop_time) <= 3 and current_frequency == frequency:
                        column+=1
                        row-=3
                    if int(loop_time) == loop and current_frequency == frequency:
                        row+=3
                        column=2
                        merge_num = int(fre_row) + 2
                        print(loop_time)
                        sheet.merge_cells(start_row=fre_row, start_column=fre_col, end_row=merge_num, end_column=fre_col) 
                        # sheet.cell(row = fre_row, column = fre_col).value = frequency
                        cell = sheet.cell(row = fre_row, column = fre_col)
                        cell.value = frequency
                        cell.border = thin_border
                        fre_row+=3
    # output excel file to current path
    wb.save(os.path.join(report_path,'output','{}.xlsx'.format(log_file_name.split(".")[0])))

parser = argparse.ArgumentParser('Auto run the specify WL case', add_help=False)
parser.add_argument("--platform", "--p", default="SPR", type=str, help="wsf run platform")

pass_args = parser.parse_args()
platform = pass_args.platform

base_path='/home/yangkun/lab/yangkunx/build-gptj/workload/GPTJ-PyTorch-Public/report'

if platform.lower() == "spr":
    report_path='{0}/{1}'.format(base_path,'spr')
elif platform.lower() == "emr":
    report_path='{0}/{1}'.format(base_path,'emr')  
else:
    print('Not support')

#create dir
check_and_create_dir(os.path.join(report_path,"output"))


# frequencys=("2.0Ghz","2.2Ghz","2.4Ghz","2.6Ghz","2.8Ghz","3.0Ghz","3.2Ghz","3.4Ghz","3.6Ghz","3.8Ghz")
frequencys=("2.8Ghz","3.0Ghz" , "3.4Ghz" ,"3.8Ghz")
loop=1
for file in os.listdir(report_path):
    if os.path.isfile(os.path.join(report_path,file)):
        parse_to_excel(platform, report_path, file, loop, frequencys)
